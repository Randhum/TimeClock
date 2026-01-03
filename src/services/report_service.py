"""
Working Time (WT) Report Generator
Generates detailed working time reports per employee for HR purposes.
"""
import datetime
import logging
from collections import deque
from typing import List, Dict, Optional
from ..data.database import Employee, TimeEntry, ensure_db_connection

logger = logging.getLogger(__name__)


def _format_hms(total_seconds: int) -> str:
    """Format seconds to HH:MM:SS."""
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


class WorkingTimeReport:
    """Generates working time reports for employees"""
    
    def __init__(self, employee: Employee, start_date: Optional[datetime.date] = None, 
                 end_date: Optional[datetime.date] = None):
        """
        Initialize a working time report for an employee.
        
        Args:
            employee: Employee to generate report for
            start_date: Start date for report (defaults to first entry date)
            end_date: End date for report (defaults to today)
        """
        self.employee = employee
        self.start_date = start_date
        self.end_date = end_date or datetime.date.today()
        self.daily_sessions = []  # List of daily work sessions
        self.total_hours = 0.0
        self.total_minutes = 0
        
    def generate(self) -> Dict:
        """
        Generate the working time report.
        
        Returns:
            Dictionary containing report data:
            {
                'employee': Employee object,
                'start_date': start_date,
                'end_date': end_date,
                'daily_sessions': List of daily session data,
                'total_hours': Total hours worked,
                'total_days': Number of days worked,
                'summary': Summary statistics
            }
        """
        ensure_db_connection()
        
        # Reset state to prevent duplicate accumulation on repeated calls
        self.daily_sessions = []
        self.total_hours = 0.0
        self.total_minutes = 0
        
        # Get all time entries for employee in date range
        entries = self._get_time_entries()
        
        if not entries:
            logger.info(f"No time entries found for {self.employee.name}")
            return self._empty_report()
        
        # If no start_date specified, use first entry date
        if not self.start_date:
            first_entry = entries[0]
            self.start_date = first_entry.timestamp.date()
        
        # Process entries into daily sessions
        self._process_entries(entries)
        
        # Calculate totals
        self._calculate_totals()
        
        return {
            'employee': self.employee,
            'start_date': self.start_date,
            'end_date': self.end_date,
            'daily_sessions': self.daily_sessions,
            'total_hours': self.total_hours,
            'total_days': len(self.daily_sessions),
            'summary': self._generate_summary()
        }
    
    def _get_time_entries(self) -> List[TimeEntry]:
        """
        Get all time entries for the employee in the date range.
        Includes clock-out entries up to 24 hours after end_date to capture
        sessions that started on the last day but ended after midnight.
        """
        query = TimeEntry.select().where(
            TimeEntry.employee == self.employee,
            TimeEntry.active == True
        ).order_by(TimeEntry.timestamp.asc())
        
        # Filter by date range if specified
        if self.start_date:
            start_datetime = datetime.datetime.combine(self.start_date, datetime.time.min)
            query = query.where(TimeEntry.timestamp >= start_datetime)
        
        # Extend end_datetime by 24 hours to include clock-out entries for sessions
        # that started on end_date but clocked out after midnight
        end_datetime = datetime.datetime.combine(self.end_date, datetime.time.max)
        end_datetime_extended = end_datetime + datetime.timedelta(days=1)
        query = query.where(TimeEntry.timestamp <= end_datetime_extended)
        
        return list(query)
    
    def _process_entries(self, entries: List[TimeEntry]):
        """
        Process time entries into daily work sessions.
        Handles sessions that span across midnight by processing all entries chronologically.
        """
        sessions = []
        pending_ins = deque()
        
        # Process all entries chronologically (already sorted by timestamp)
        for entry in entries:
            if entry.action == 'in':
                pending_ins.append((entry.timestamp, entry.id))
            elif entry.action == 'out':
                if not pending_ins:
                    logger.warning(f"Clock out without clock in for {self.employee.name} on {entry.timestamp.date()}")
                    continue
                clock_in_time, clock_in_id = pending_ins.popleft()
                duration = entry.timestamp - clock_in_time
                total_seconds = int(duration.total_seconds())
                
                # Use clock_in date for assigning session to a day
                # This ensures sessions spanning midnight are counted on the day they started
                session_date = clock_in_time.date()
                
                sessions.append({
                    'date': session_date,
                    'clock_in': clock_in_time,
                    'clock_out': entry.timestamp,
                    'total_seconds': total_seconds,
                    'total_minutes': total_seconds // 60,
                    'formatted_time': _format_hms(total_seconds),
                    'clock_in_entry_id': clock_in_id,
                    'clock_out_entry_id': entry.id
                })
        
        if pending_ins:
            logger.info(f"{len(pending_ins)} open session(s) for {self.employee.name} remain without a clock-out")
        
        # Convert sessions to daily_sessions format
        for session in sessions:
            total_seconds = session.get('total_seconds', session['total_minutes'] * 60)
            hours = total_seconds / 3600.0
            minutes = total_seconds / 60.0
            
            self.daily_sessions.append({
                'date': session['date'],
                'clock_in': session['clock_in'],
                'clock_out': session['clock_out'],
                'hours': hours,
                'minutes': minutes,
                'total_minutes': session['total_minutes'],
                'total_seconds': total_seconds,
                'formatted_time': session['formatted_time']
            })
    
    def _calculate_totals(self):
        """Calculate total hours/minutes/seconds worked"""
        total_seconds = sum(session.get('total_seconds', session['total_minutes'] * 60) for session in self.daily_sessions)
        self.total_hours = total_seconds / 3600.0
        self.total_minutes = total_seconds / 60.0
        self.total_seconds = total_seconds
    
    def _generate_summary(self) -> Dict:
        """Generate summary statistics"""
        if not self.daily_sessions:
            return {
                'total_hours': 0,
                'total_minutes': 0,
                'formatted_total': "00:00:00",
                'average_hours_per_day': 0,
                'formatted_average_per_day': "00:00:00",
                'days_worked': 0
            }
        
        days_worked = len(set(session['date'] for session in self.daily_sessions))
        average_seconds = int(round(self.total_seconds / days_worked)) if days_worked else 0
        avg_hours = average_seconds / 3600.0

        total_hours_int = self.total_seconds // 3600
        total_mins_int = (self.total_seconds % 3600) // 60
        total_secs_int = self.total_seconds % 60
        
        return {
            'total_hours': self.total_hours,
            'total_minutes': self.total_minutes,
            'total_seconds': self.total_seconds,
            'formatted_total': f"{int(total_hours_int):02d}:{int(total_mins_int):02d}:{int(total_secs_int):02d}",
            'average_hours_per_day': avg_hours,
            'formatted_average_per_day': _format_hms(average_seconds),
            'days_worked': days_worked
        }
    
    def _empty_report(self) -> Dict:
        """Return empty report structure"""
        return {
            'employee': self.employee,
            'start_date': self.start_date or datetime.date.today(),
            'end_date': self.end_date,
            'daily_sessions': [],
            'total_hours': 0.0,
            'total_days': 0,
            'summary': {
                'total_hours': 0,
                'total_minutes': 0,
                'total_seconds': 0,
                'formatted_total': "00:00:00",
                'average_hours_per_day': 0,
                'formatted_average_per_day': "00:00:00",
                'days_worked': 0
            }
        }
    
    def to_csv(
        self,
        filename: Optional[str] = None,
        export_root: Optional[str] = None
    ) -> str:
        """
        Export report to CSV file.
        
        Args:
            filename: Optional filename. If not provided, generates one.
            export_root: Optional directory where the file should be written.
        
        Returns:
            Path to the generated file.
        """
        import csv
        import io
        import os
        
        report = self.generate()
        
        if filename is None:
            start_str = report['start_date'].strftime('%Y%m%d')
            end_str = report['end_date'].strftime('%Y%m%d')
            safe_name = "".join(c for c in report['employee'].name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')
            root = export_root or os.path.join(os.getcwd(), 'exports')
            filename = os.path.join(root, f"WT_Report_{safe_name}_{start_str}_{end_str}.csv")
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(['Working Time Report'])
        writer.writerow(['Employee:', report['employee'].name])
        writer.writerow(['Employee ID:', report['employee'].rfid_tag])
        writer.writerow(['Period:', f"{report['start_date']} to {report['end_date']}"])
        writer.writerow([])
        
        # Daily sessions
        writer.writerow(['Date', 'Clock In', 'Clock Out', 'Hours Worked (HH:MM:SS)'])
        for session in report['daily_sessions']:
            writer.writerow([
                session['date'].strftime('%Y-%m-%d'),
                session['clock_in'].strftime('%H:%M:%S'),
                session['clock_out'].strftime('%H:%M:%S'),
                session['formatted_time']
            ])
        
        writer.writerow([])
        
        # Summary
        summary = report['summary']
        writer.writerow(['Summary'])
        writer.writerow(['Total Hours:', f"{summary['formatted_total']}"])
        writer.writerow(['Days Worked:', summary['days_worked']])
        writer.writerow(['Average Hours per Day:', summary['formatted_average_per_day']])

        # Write plain-text CSV
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else 'exports', exist_ok=True)
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            csvfile.write(buffer.getvalue())
        logger.info(f"WT Report export written to {filename}")

        return filename
    
    def to_text(self) -> str:
        """
        Generate a human-readable text report.
        
        Returns:
            Formatted text report
        """
        report = self.generate()
        lines = []
        
        lines.append("=" * 37)
        lines.append("WORKING TIME REPORT")
        lines.append("=" * 37)
        lines.append(f"Name: {report['employee'].name}")
        
        if not report['daily_sessions']:
            lines.append("No time entries found for this period.")
            return "\n".join(lines)
        
        # Daily sessions
        lines.append("-" * 37)
        lines.append(f"{'Date':<12} {'Clock In':<12} {'Clock Out':<12} {'Hours':<10}")
        lines.append("-" * 37)
        
        for session in report['daily_sessions']:
            lines.append(
                f"{session['date'].strftime('%Y-%m-%d'):<12} "
                f"{session['clock_in'].strftime('%H:%M:%S'):<12} "
                f"{session['clock_out'].strftime('%H:%M:%S'):<12} "
                f"{session['formatted_time']:<10}"
            )
        
        lines.append("-" * 37)
        
        # Summary
        summary = report['summary']
        lines.append("SUMMARY")
        lines.append("-" * 37)
        lines.append(f"Total Hours Worked: {summary['formatted_total']}")
        lines.append(f"Days Worked: {summary['days_worked']}")
        lines.append(f"Average Hours per Day: {summary['formatted_average_per_day']}")
        lines.append("=" * 37)
        
        return "\n".join(lines)
    
    def _build_lgav_data(self) -> Dict:
        """
        Build simple monthly hours data structure.
        
        Returns:
            Dictionary with hours worked per day, organized by month.
        """
        report = self.generate()
        start = report['start_date']
        end = report['end_date']
        
        # Create a map of date -> total seconds worked that day (from clock entries)
        # Only include sessions that started within the date range
        daily_totals = {}
        for session in report['daily_sessions']:
            date = session['date']
            # Only include sessions that started within the report date range
            if start <= date <= end:
                if date not in daily_totals:
                    daily_totals[date] = 0
                daily_totals[date] += session['total_seconds']
        
        # Organize by months
        months = {}
        current_date = start
        
        while current_date <= end:
            year_month = (current_date.year, current_date.month)
            if year_month not in months:
                months[year_month] = {
                    'year': current_date.year,
                    'month': current_date.month,
                    'days': {},
                    'total_seconds': 0,
                }
            
            day_of_month = current_date.day
            month_data = months[year_month]
            
            # Get hours worked from clock entries
            total_seconds = daily_totals.get(current_date, 0)
            
            month_data['days'][day_of_month] = {
                'date': current_date,
                'total_seconds': total_seconds
            }
            month_data['total_seconds'] += total_seconds
            
            current_date += datetime.timedelta(days=1)
        
        return {
            'employee': report['employee'],
            'start_date': start,
            'end_date': end,
            'months': months
        }
    
    def to_lgav_excel(
        self,
        filename: Optional[str] = None,
        export_root: Optional[str] = None
    ) -> str:
        """
        Export working hours to Excel - simple format with hours per day per month.
        
        Args:
            filename: Optional filename. If not provided, generates one.
            export_root: Optional directory where the file should be written.
        
        Returns:
            Path to the generated file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError(f"Required libraries not installed: {e}. Install with: pip install openpyxl")
        
        import os
        import calendar
        
        lgav_data = self._build_lgav_data()
        report = self.generate()
        
        if filename is None:
            start_str = report['start_date'].strftime('%Y%m%d')
            end_str = report['end_date'].strftime('%Y%m%d')
            safe_name = "".join(c for c in report['employee'].name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')
            root = export_root or os.path.join(os.getcwd(), 'exports')
            filename = os.path.join(root, f"Arbeitszeit_{safe_name}_{start_str}_{end_str}.xlsx")
        
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else 'exports', exist_ok=True)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Arbeitszeit"
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        
        # Month names in German
        month_names = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                       'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
        
        # ===== ROW 1: Employee Name =====
        ws['A1'] = f"Arbeitszeitnachweis: {report['employee'].name}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:AF1')
        
        # ===== ROW 2: Date Range =====
        ws['A2'] = f"Zeitraum: {report['start_date'].strftime('%d.%m.%Y')} - {report['end_date'].strftime('%d.%m.%Y')}"
        ws['A2'].font = normal_font
        ws.merge_cells('A2:AF2')
        
        row = 4  # Start data from row 4
        
        # Process each month
        for (year, month), month_data in sorted(lgav_data['months'].items()):
            days_in_month = calendar.monthrange(year, month)[1]
            
            # Month header row
            ws[f'A{row}'] = f"{month_names[month]} {year}"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:AF{row}')
            row += 1
            
            # Day numbers row
            for day in range(1, days_in_month + 1):
                col = get_column_letter(day)
                cell = ws[f'{col}{row}']
                cell.value = day
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
            
            # Total column header
            total_col = get_column_letter(days_in_month + 1)
            ws[f'{total_col}{row}'] = "Total"
            ws[f'{total_col}{row}'].font = header_font
            ws[f'{total_col}{row}'].alignment = center_align
            ws[f'{total_col}{row}'].border = thin_border
            ws[f'{total_col}{row}'].fill = header_fill
            row += 1
            
            # Hours row
            month_total_seconds = 0
            for day in range(1, days_in_month + 1):
                col = get_column_letter(day)
                cell = ws[f'{col}{row}']
                
                if day in month_data['days']:
                    total_seconds = month_data['days'][day]['total_seconds']
                    if total_seconds > 0:
                        # Format as H:MM
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        cell.value = f"{hours}:{minutes:02d}"
                        month_total_seconds += total_seconds
                    else:
                        cell.value = ""
                else:
                    cell.value = ""
                
                cell.alignment = center_align
                cell.border = thin_border
            
            # Month total
            total_hours = month_total_seconds // 3600
            total_minutes = (month_total_seconds % 3600) // 60
            ws[f'{total_col}{row}'] = f"{total_hours}:{total_minutes:02d}"
            ws[f'{total_col}{row}'].font = header_font
            ws[f'{total_col}{row}'].alignment = center_align
            ws[f'{total_col}{row}'].border = thin_border
            
            row += 2  # Empty row between months
        
        # Column widths
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 6
        
        wb.save(filename)
        logger.info(f"Working time Excel report exported to {filename}")
        return filename
    
    def to_lgav_csv(
        self,
        filename: Optional[str] = None,
        export_root: Optional[str] = None
    ) -> str:
        """
        Export working hours to CSV - simple format with hours per day per month.
        
        Args:
            filename: Optional filename. If not provided, generates one.
            export_root: Optional directory where the file should be written.
        
        Returns:
            Path to the generated file.
        """
        import csv
        import os
        import calendar
        
        lgav_data = self._build_lgav_data()
        report = self.generate()
        
        # Month names in German
        month_names = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                       'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
        
        if filename is None:
            start_str = report['start_date'].strftime('%Y%m%d')
            end_str = report['end_date'].strftime('%Y%m%d')
            safe_name = "".join(c for c in report['employee'].name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')
            root = export_root or os.path.join(os.getcwd(), 'exports')
            filename = os.path.join(root, f"Arbeitszeit_{safe_name}_{start_str}_{end_str}.csv")
        
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else 'exports', exist_ok=True)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')  # Semicolon as per European CSV standard
            
            # Header
            writer.writerow([f'Arbeitszeitnachweis: {report["employee"].name}'])
            writer.writerow([f'Zeitraum: {report["start_date"].strftime("%d.%m.%Y")} - {report["end_date"].strftime("%d.%m.%Y")}'])
            writer.writerow([])
            
            # Process each month
            for (year, month), month_data in sorted(lgav_data['months'].items()):
                days_in_month = calendar.monthrange(year, month)[1]
                
                # Month header
                writer.writerow([f'{month_names[month]} {year}'])
                
                # Day numbers row
                day_row = [str(day) for day in range(1, days_in_month + 1)] + ['Total']
                writer.writerow(day_row)
                
                # Hours row
                hours_row = []
                month_total_seconds = 0
                for day in range(1, days_in_month + 1):
                    if day in month_data['days']:
                        total_seconds = month_data['days'][day]['total_seconds']
                        if total_seconds > 0:
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            hours_row.append(f"{hours}:{minutes:02d}")
                            month_total_seconds += total_seconds
                        else:
                            hours_row.append('')
                    else:
                        hours_row.append('')
                
                # Month total
                total_hours = month_total_seconds // 3600
                total_minutes = (month_total_seconds % 3600) // 60
                hours_row.append(f"{total_hours}:{total_minutes:02d}")
                writer.writerow(hours_row)
                
                writer.writerow([])  # Empty row between months
        
        logger.info(f"Working time CSV report exported to {filename}")
        return filename
    
    def to_lgav_pdf(
        self,
        filename: Optional[str] = None,
        export_root: Optional[str] = None
    ) -> str:
        """
        Export working hours to PDF - simple format with hours per day per month.
        
        Args:
            filename: Optional filename. If not provided, generates one.
            export_root: Optional directory where the file should be written.
        
        Returns:
            Path to the generated file.
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm, mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError as e:
            raise ImportError(f"Required libraries not installed: {e}. Install with: pip install reportlab")
        
        import os
        import calendar
        
        lgav_data = self._build_lgav_data()
        report = self.generate()
        
        # Month names in German
        month_names = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                       'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
        
        if filename is None:
            start_str = report['start_date'].strftime('%Y%m%d')
            end_str = report['end_date'].strftime('%Y%m%d')
            safe_name = "".join(c for c in report['employee'].name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')
            root = export_root or os.path.join(os.getcwd(), 'exports')
            filename = os.path.join(root, f"Arbeitszeit_{safe_name}_{start_str}_{end_str}.pdf")
        
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else 'exports', exist_ok=True)
        
        # Create PDF in landscape for wide table
        doc = SimpleDocTemplate(filename, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#000000'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        story.append(Paragraph(f"Arbeitszeitnachweis: {report['employee'].name}", title_style))
        
        # Date range
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"Zeitraum: {report['start_date'].strftime('%d.%m.%Y')} - {report['end_date'].strftime('%d.%m.%Y')}",
            info_style
        ))
        story.append(Spacer(1, 0.5*cm))
        
        # Process each month
        grand_total_seconds = 0
        
        for (year, month), month_data in sorted(lgav_data['months'].items()):
            days_in_month = calendar.monthrange(year, month)[1]
            
            # Month header
            month_style = ParagraphStyle(
                'MonthHeader',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#333333'),
                spaceBefore=10,
                spaceAfter=6
            )
            story.append(Paragraph(f"{month_names[month]} {year}", month_style))
            
            # Build table: day numbers row + hours row
            day_row = [str(d) for d in range(1, days_in_month + 1)] + ['Total']
            hours_row = []
            month_total_seconds = 0
            
            for day in range(1, days_in_month + 1):
                if day in month_data['days']:
                    total_seconds = month_data['days'][day]['total_seconds']
                    if total_seconds > 0:
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        hours_row.append(f"{hours}:{minutes:02d}")
                        month_total_seconds += total_seconds
                    else:
                        hours_row.append('')
                else:
                    hours_row.append('')
            
            # Month total
            total_hours = month_total_seconds // 3600
            total_minutes = (month_total_seconds % 3600) // 60
            hours_row.append(f"{total_hours}:{total_minutes:02d}")
            grand_total_seconds += month_total_seconds
            
            # Create table
            table_data = [day_row, hours_row]
            col_width = (landscape(A4)[0] - 30*mm) / (days_in_month + 1)
            table = Table(table_data, colWidths=[col_width] * (days_in_month + 1))
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCFF')),
                ('BACKGROUND', (-1, 1), (-1, 1), colors.HexColor('#E8E8FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (-1, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.3*cm))
        
        # Grand total
        story.append(Spacer(1, 0.3*cm))
        grand_hours = grand_total_seconds // 3600
        grand_minutes = (grand_total_seconds % 3600) // 60
        total_style = ParagraphStyle(
            'GrandTotal',
            parent=styles['Heading2'],
            fontSize=12,
            alignment=TA_LEFT
        )
        story.append(Paragraph(f"<b>Gesamtstunden:</b> {grand_hours}:{grand_minutes:02d}", total_style))
        
        doc.build(story)
        logger.info(f"Working time PDF report exported to {filename}")
        return filename


def generate_wt_report(employee: Employee, start_date: Optional[datetime.date] = None,
                      end_date: Optional[datetime.date] = None) -> WorkingTimeReport:
    """
    Convenience function to create and generate a WT report.
    
    Args:
        employee: Employee to generate report for
        start_date: Start date (optional)
        end_date: End date (optional, defaults to today)
    
    Returns:
        WorkingTimeReport object with generated report
    """
    report = WorkingTimeReport(employee, start_date, end_date)
    report.generate()
    return report


def generate_all_employees_lgav_excel(
    filename: Optional[str] = None,
    export_root: Optional[str] = None,
    start_date: Optional[datetime.date] = None,
    end_date: Optional[datetime.date] = None
) -> str:
    """
    Generate LGAV Excel report for all active employees with one sheet per employee.
    
    Args:
        filename: Optional filename. If not provided, generates one.
        export_root: Optional directory where the file should be written.
        start_date: Start date for report (defaults to one year ago)
        end_date: End date for report (defaults to today)
    
    Returns:
        Path to the generated file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(f"Required libraries not installed: {e}. Install with: pip install openpyxl")
    
    import os
    import calendar
    from ..data.database import get_all_employees
    
    ensure_db_connection()
    
    # Set default dates: last year from today
    if end_date is None:
        end_date = datetime.date.today()
    if start_date is None:
        start_date = end_date - datetime.timedelta(days=365)
    
    # Get all active employees
    employees = list(get_all_employees(include_inactive=False))
    
    if not employees:
        raise ValueError("No active employees found")
    
    # Generate filename if not provided
    if filename is None:
        start_str = start_date.strftime('%Y%m%d')
        end_str = end_date.strftime('%Y%m%d')
        root = export_root or os.path.join(os.getcwd(), 'exports')
        filename = os.path.join(root, f"LGAV_Alle_Mitarbeiter_{start_str}_{end_str}.xlsx")
    
    os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else 'exports', exist_ok=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Month names in German
    month_names = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                   'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
    
    # Process each employee
    for employee in employees:
        # Create report for this employee
        report = WorkingTimeReport(employee, start_date, end_date)
        lgav_data = report._build_lgav_data()
        
        # Create worksheet for this employee (sheet name max 31 chars)
        safe_sheet_name = "".join(c for c in employee.name if c.isalnum() or c in (' ', '-', '_'))[:31]
        if not safe_sheet_name:
            safe_sheet_name = f"Employee_{employee.id}"
        ws = wb.create_sheet(title=safe_sheet_name)
        
        # ===== ROW 1: Employee Name =====
        ws['A1'] = f"Arbeitszeitnachweis: {employee.name}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:AF1')
        
        # ===== ROW 2: Date Range =====
        ws['A2'] = f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws['A2'].font = normal_font
        ws.merge_cells('A2:AF2')
        
        row = 4  # Start data from row 4
        
        # Process each month
        for (year, month), month_data in sorted(lgav_data['months'].items()):
            days_in_month = calendar.monthrange(year, month)[1]
            
            # Month header row
            ws[f'A{row}'] = f"{month_names[month]} {year}"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:AF{row}')
            row += 1
            
            # Day numbers row
            for day in range(1, days_in_month + 1):
                col = get_column_letter(day)
                cell = ws[f'{col}{row}']
                cell.value = day
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
            
            # Total column header
            total_col = get_column_letter(days_in_month + 1)
            ws[f'{total_col}{row}'] = "Total"
            ws[f'{total_col}{row}'].font = header_font
            ws[f'{total_col}{row}'].alignment = center_align
            ws[f'{total_col}{row}'].border = thin_border
            ws[f'{total_col}{row}'].fill = header_fill
            row += 1
            
            # Hours row
            month_total_seconds = 0
            for day in range(1, days_in_month + 1):
                col = get_column_letter(day)
                cell = ws[f'{col}{row}']
                
                if day in month_data['days']:
                    total_seconds = month_data['days'][day]['total_seconds']
                    if total_seconds > 0:
                        # Format as H:MM
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        cell.value = f"{hours}:{minutes:02d}"
                        month_total_seconds += total_seconds
                    else:
                        cell.value = ""
                else:
                    cell.value = ""
                
                cell.alignment = center_align
                cell.border = thin_border
            
            # Month total
            total_hours = month_total_seconds // 3600
            total_minutes = (month_total_seconds % 3600) // 60
            ws[f'{total_col}{row}'] = f"{total_hours}:{total_minutes:02d}"
            ws[f'{total_col}{row}'].font = header_font
            ws[f'{total_col}{row}'].alignment = center_align
            ws[f'{total_col}{row}'].border = thin_border
            
            row += 2  # Empty row between months
        
        # Column widths
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 6
    
    wb.save(filename)
    logger.info(f"LGAV Excel report for all employees exported to {filename}")
    return filename
